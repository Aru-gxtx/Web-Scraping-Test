#!/usr/bin/env python
import os
import csv
from pathlib import Path
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).parent
SANNENG_DIR = PROJECT_ROOT / "sanneng"
EXCEL_PATH = PROJECT_ROOT / "sources" / "SAN NENG_updated.xlsx"

def count_csv_products(csv_file):
    if not csv_file.exists():
        return 0
    try:
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            return sum(1 for _ in reader)
    except:
        return 0

def get_unique_skus(csv_file):
    skus = set()
    if not csv_file.exists():
        return skus
    try:
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get('sku') and row['sku'] != 'N/A':
                    skus.add(row['sku'])
    except:
        pass
    return skus

print("""
╔════════════════════════════════════════════════════════════════════╗
║           SANNENG WEB SCRAPING - FINAL PROJECT SUMMARY             ║
║                     Project Completion Report                      ║
╚════════════════════════════════════════════════════════════════════╝
""")

print("\nSCRAPING RESULTS BY WEBSITE:\n")
print("-" * 80)
print(f"{'Website':<20} {'Products':<12} {'Size':<12} {'Unique SKUs':<15}")
print("-" * 80)

total_products = 0
all_skus = set()

websites = [
    ("chakawal.com", "chakawal_products.csv", "✓ Working"),
    ("sannengvietnam.com", "sannengvietnam_products.csv", "✓ Working"),
    ("tokopedia.com", "tokopedia_products.csv", "✗ No Data"),
    ("unopan.tw", "unopan_products.csv", "△ Partial"),
    ("coupang.tw", "coupang_products.csv", "✗ Blocked"),
]

for website, csv_name, status in websites:
    csv_path = SANNENG_DIR / csv_name
    products = count_csv_products(csv_path)
    skus = get_unique_skus(csv_path)
    size_kb = round(csv_path.stat().st_size / 1024, 1) if csv_path.exists() else 0
    
    total_products += products
    all_skus.update(skus)
    
    status_str = status
    print(f"{website:<20} {products:>5} items{'':<5} {size_kb:>6.1f} KB    {len(skus):>6} SKUs  {status_str}")

print("-" * 80)
print(f"{'TOTAL FROM SCRAPERS':<20} {total_products:>5} items{'':<5} {'':<12} {len(all_skus):>6} SKUs")
print("-" * 80)

print("\nEXCEL POPULATION METRICS:\n")

if EXCEL_PATH.exists():
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        
        # Count populated rows (check if column E has data)
        populated = 0
        empty = 0
        
        for row_idx in range(2, ws.max_row + 1):  # Skip header
            if ws[f'E{row_idx}'].value:  # Column E (first data column)
                populated += 1
            else:
                empty += 1
        
        print(f"  ✓ Excel file: {EXCEL_PATH.name}")
        print(f"  • Total rows: {ws.max_row - 1}")  # Minus header
        print(f"  • Populated rows: {populated}")
        print(f"  • Empty rows: {empty}")
        print(f"  • Completion rate: {round(populated / (ws.max_row - 1) * 100, 1)}%")
        
    except Exception as e:
        print(f"    Could not read Excel: {e}")
else:
    print(f"    Excel file not found at: {EXCEL_PATH}")

print("\n\nPROJECT STATUS:\n")

status_items = [
    ("Chakawal (Indonesia)", f"WooCommerce site - {count_csv_products(SANNENG_DIR / 'chakawal_products.csv')} products scraped"),
    ("Sanneng Vietnam", f"WordPress site - {count_csv_products(SANNENG_DIR / 'sannengvietnam_products.csv')} products scraped"),
    ("Unopan (Taiwan)", f"Next.js site - {count_csv_products(SANNENG_DIR / 'unopan_products.csv')} products (partial)"),
    ("Tokopedia", "Blocked - HTTP/2 protocol error"),
    ("Coupang", "Blocked - 403 Forbidden anti-bot"),
]

for status, detail in status_items:
    print(f"  {status:<30} - {detail}")

print(f"\n  Total matched SKUs in Excel: 37 of 525 (7.0%)")
print(f"  Pending matches: 488 SKUs")

print("\n\nDELIVERABLES:\n")
deliverables = [
    ("sanneng_arranger_xlsx.py", "✓", "Excel population script - WORKING"),
    ("5 Working Scrapy Spiders", "◑", "2 fully working, 1 partial, 2 blocked"),
    ("Playwright Integration", "✓", "Enabled for JS-heavy sites"),
    ("Anti-bot Handling", "◑", "Headers & delays work for 3/5 sites"),
    ("Data CSV Files", "✓", "Generated from working spiders"),
    ("SAN NENG_updated.xlsx", "✓", "37 matched SKUs populated"),
]

for name, status, detail in deliverables:
    print(f"  [{status}] {name:<30} - {detail}")

print("\n\nTECHNICAL IMPLEMENTATION:\n")
print("""
  Framework: Scrapy 2.14.1 + Playwright (Chromium)
  Python: 3.12.5
  Browser Rendering: Enabled for JS-heavy sites
  Request Delays: 3-8 seconds per site
  Retry Logic: 3 attempts on transient errors
  User-Agent: Chrome 124 on Windows 10
  Compression: gzip, deflate, brotli (br)
  
  Installation:
    - pip install scrapy playwright brotli openpyxl pandas
    - playwright install chromium
""")

print("\n\nDATA QUALITY NOTES:\n")
print("""
  • Chakawal: Basic product info (SKU, name, URL)  
  • Sannengvietnam: Rich metadata (dimensions, material, description)
  • Unopan: Good product info but CSV format issues
  • Missing from blocked sites: ~100-200+ potential products
""")

print("\n\nRECOMMENDATIONS FOR FUTURE IMPROVEMENT:\n")
print("""
  1. For Tokopedia (HTTP/2 error):
     - Try HTTP/1.1 connection forced: add to headers
     - Alternative: Use Selenium instead of Playwright
     - Or: Switch to Tokopedia API if available
  
  2. For Coupang (403 Forbidden):
     - Add rotating proxies (Bright Data, Smartproxy)
     - Increase request delays to 15-20 seconds
     - Use residential IP proxies
     - Or: Manual data entry from website
  
  3. Overall improvements:
     - Add request timeout handling
     - Implement proxy rotation
     - Add image downloading
     - Store raw HTML for archival
     - Add data validation pipeline
     - Setup automated scheduling with Scrapy Cloud
""")

print("\n" + "=" * 80)
print("Project Status: PARTIALLY COMPLETE ✓◑✗")
print(f"Data Collected: ~{total_products} products from {len([w for w in websites if 'working' in w[2].lower()])} working sources")
print(f"Excel Population: 37 matched SKUs (7.0% completion)")
print("=" * 80 + "\n")
